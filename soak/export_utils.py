"""Utility functions for exporting data from DAG nodes."""

import json
import logging
from pathlib import Path
from typing import Dict, List, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

logger = logging.getLogger(__name__)


def export_to_csv(df: pd.DataFrame, path: Path) -> None:
    """Export DataFrame to CSV file.

    Args:
        df: DataFrame to export
        path: Output path (should have .csv extension)
    """
    try:
        df.to_csv(path, index=False)
        logger.debug(f"Exported CSV to {path}")
    except Exception as e:
        logger.error(f"Failed to export CSV to {path}: {e}")
        raise


def export_to_html(df: pd.DataFrame, path: Path) -> None:
    """Export DataFrame to HTML file with styling.

    Args:
        df: DataFrame to export
        path: Output path (should have .html extension)
    """
    try:
        html = df.to_html(index=False, escape=False, na_rep="")
        styled_html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        table {{ border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; }}
        th {{ background-color: #4CAF50; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
        tr:hover {{ background-color: #f5f5f5; }}
    </style>
</head>
<body>
    {html}
</body>
</html>
"""
        path.write_text(styled_html)
        logger.debug(f"Exported HTML to {path}")
    except Exception as e:
        logger.error(f"Failed to export HTML to {path}: {e}")
        raise


def export_to_json(rows: List[Dict], path: Path) -> None:
    """Export list of dictionaries to JSON file.

    Args:
        rows: List of row dictionaries to export
        path: Output path (should have .json extension)
    """
    try:
        with open(path, "w") as f:
            json.dump(rows, f, indent=2, default=str)
        logger.debug(f"Exported JSON to {path}")
    except Exception as e:
        logger.error(f"Failed to export JSON to {path}: {e}")
        raise


def identify_text_columns(df: pd.DataFrame, threshold: int = 50) -> Set[str]:
    """Identify columns that contain long text.

    Args:
        df: DataFrame to analyze
        threshold: Character length threshold for "long text"

    Returns:
        Set of column names that contain long text
    """
    text_columns = set()
    for col in df.columns:
        if df[col].dtype == object:  # String columns
            max_len = df[col].astype(str).str.len().max()
            if max_len > threshold:
                text_columns.add(col)
    return text_columns


def apply_excel_formatting(
    file_path: Path,
    text_columns: Set[str] = None,
    font_size: int = 12,
    text_col_width: int = 60,
    header_bold: bool = True,
) -> None:
    """Apply formatting to an Excel file.

    Args:
        file_path: Path to XLSX file
        text_columns: Column names to apply text wrapping (auto-detected if None)
        font_size: Font size for all cells
        text_col_width: Width for text columns
        header_bold: Make header row bold
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # Auto-detect text columns if not provided
    if text_columns is None:
        # Read the file to detect
        df = pd.read_excel(file_path, engine="openpyxl")
        text_columns = identify_text_columns(df)

    font = Font(size=font_size)

    # Apply formatting to all cells
    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        for col_idx, cell in enumerate(row, start=1):
            cell.font = font

            # Get column name from header row
            col_name = ws.cell(1, col_idx).value

            # Apply text wrapping to long-text columns
            if col_name in text_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    # Set column widths
    for col_idx, col in enumerate(ws.columns, start=1):
        col_letter = ws.cell(1, col_idx).column_letter
        col_name = ws.cell(1, col_idx).value

        if col_name in text_columns:
            # Wide columns for long text
            ws.column_dimensions[col_letter].width = text_col_width
        else:
            # Auto-size other columns with reasonable limits
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 15), 40)
            ws.column_dimensions[col_letter].width = adjusted_width

    # Make header row bold
    if header_bold:
        for cell in ws[1]:
            cell.font = Font(size=font_size, bold=True)

    wb.save(file_path)
    logger.debug(f"Applied formatting to {file_path}")


def export_to_excel(
    df: pd.DataFrame,
    path: Path,
    font_size: int = 12,
    text_col_width: int = 60,
    header_bold: bool = True,
) -> None:
    """Export DataFrame to formatted Excel file.

    Args:
        df: DataFrame to export
        path: Output path (should have .xlsx extension)
        font_size: Font size for all cells
        text_col_width: Width for text columns
        header_bold: Make header row bold
    """
    try:
        # Write to Excel with pandas
        df.to_excel(path, index=False, engine="openpyxl")

        # Identify text columns
        text_columns = identify_text_columns(df)

        # Apply formatting
        apply_excel_formatting(
            path,
            text_columns=text_columns,
            font_size=font_size,
            text_col_width=text_col_width,
            header_bold=header_bold,
        )

        logger.debug(f"Exported formatted Excel to {path}")
    except Exception as e:
        logger.error(f"Failed to export Excel to {path}: {e}")
        raise
