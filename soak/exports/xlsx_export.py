"""XLSX export functionality for qualitative analysis results."""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def generate_self_similarity_xlsx(
    themes: List[Dict[str, Any]],
    output_path: Path,
    embedding_model: str = "text-embedding-3-small",
) -> bool:
    """Export theme self-similarity matrix to XLSX file.

    Args:
        themes: List of theme dicts with name, description, codes
        output_path: Path to write XLSX file
        embedding_model: Model to use for embeddings

    Returns:
        True if export succeeded, False if insufficient themes.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    if len(themes) < 2:
        return False

    try:
        from ..analysis.self_similarity import compute_self_similarity_matrix

        self_sim = compute_self_similarity_matrix(
            themes, embedding_model=embedding_model
        )
    except Exception as e:
        logger.warning(f"Could not compute self-similarity for XLSX: {e}")
        return False

    # colour scale matching web app results.html getCellStyle()
    # < 0.5: red, 0.5-0.7: yellow/gold, >= 0.7: green
    fill_green = PatternFill(
        start_color="FF50C85A", end_color="FF50C85A", fill_type="solid"
    )  # rgb(80,200,90)
    fill_yellow = PatternFill(
        start_color="FFF0C85A", end_color="FFF0C85A", fill_type="solid"
    )  # rgb(240,200,90)
    fill_red = PatternFill(
        start_color="FFC85050", end_color="FFC85050", fill_type="solid"
    )  # rgb(200,80,80)
    fill_diagonal = PatternFill(
        start_color="FFE5E5E5", end_color="FFE5E5E5", fill_type="solid"
    )

    font_white = Font(color="FFFFFFFF")
    font_dark = Font(color="FF333333")
    font_grey = Font(color="FF999999")

    def get_fill_and_font(val):
        """Return fill and font for a similarity value."""
        if val >= 0.7:
            return fill_green, font_dark
        elif val >= 0.5:
            return fill_yellow, font_dark
        else:
            return fill_red, font_white

    wb = openpyxl.Workbook()

    # sheet 1: calibrated matrix (upper triangle only)
    ws_cal = wb.active
    ws_cal.title = "Calibrated"

    # header row
    ws_cal.cell(row=1, column=1, value="Theme")
    for i, label in enumerate(self_sim["short_labels"]):
        cell = ws_cal.cell(row=1, column=i + 2, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # data rows - upper triangle only
    for i, theme_name in enumerate(self_sim["themes"]):
        ws_cal.cell(
            row=i + 2, column=1, value=f"{self_sim['short_labels'][i]}: {theme_name}"
        )
        ws_cal.cell(row=i + 2, column=1).font = Font(bold=True)

        for j in range(self_sim["n_themes"]):
            cell = ws_cal.cell(row=i + 2, column=j + 2)
            cell.alignment = Alignment(horizontal="center")

            if i == j:
                # diagonal
                cell.value = "-"
                cell.fill = fill_diagonal
                cell.font = font_grey
            elif j > i:
                # upper triangle - show value with colour
                val = self_sim["calibrated_matrix"][i][j]
                cell.value = round(val, 2)
                fill, font = get_fill_and_font(val)
                cell.fill = fill
                cell.font = font
            # lower triangle: leave empty

    # auto-width first column
    ws_cal.column_dimensions["A"].width = 40

    # sheet 2: raw matrix (upper triangle only)
    ws_raw = wb.create_sheet(title="Raw (Cosine)")

    ws_raw.cell(row=1, column=1, value="Theme")
    for i, label in enumerate(self_sim["short_labels"]):
        cell = ws_raw.cell(row=1, column=i + 2, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, theme_name in enumerate(self_sim["themes"]):
        ws_raw.cell(
            row=i + 2, column=1, value=f"{self_sim['short_labels'][i]}: {theme_name}"
        )
        ws_raw.cell(row=i + 2, column=1).font = Font(bold=True)

        for j in range(self_sim["n_themes"]):
            cell = ws_raw.cell(row=i + 2, column=j + 2)
            cell.alignment = Alignment(horizontal="center")

            if i == j:
                cell.value = "-"
                cell.fill = fill_diagonal
                cell.font = font_grey
            elif j > i:
                val = self_sim["raw_matrix"][i][j]
                cell.value = round(val, 3)
            # lower triangle: leave empty

    ws_raw.column_dimensions["A"].width = 40

    # sheet 3: legend
    ws_legend = wb.create_sheet(title="Legend")
    ws_legend.cell(row=1, column=1, value="Theme Label").font = Font(bold=True)
    ws_legend.cell(row=1, column=2, value="Theme Name").font = Font(bold=True)

    for i, theme_name in enumerate(self_sim["themes"]):
        ws_legend.cell(row=i + 2, column=1, value=self_sim["short_labels"][i])
        ws_legend.cell(row=i + 2, column=2, value=theme_name)

    ws_legend.column_dimensions["A"].width = 12
    ws_legend.column_dimensions["B"].width = 50

    wb.save(output_path)
    return True


def generate_themes_xlsx(
    themes: List[Dict[str, Any]],
    codes: List[Dict[str, Any]],
    output_path: Path,
) -> None:
    """Export themes to XLSX file.

    Args:
        themes: List of theme dicts
        codes: List of code dicts (for extracting quotes)
        output_path: Path to write XLSX file
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Themes"

    # header
    headers = ["#", "Name", "Description", "Code Count", "Codes", "Sample Quotes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # data
    for row, theme in enumerate(themes, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=theme.get("name", ""))
        ws.cell(row=row, column=3, value=theme.get("description", ""))

        code_list = theme.get(
            "codes", theme.get("code_hashes", theme.get("code_slugs", []))
        )
        ws.cell(row=row, column=4, value=len(code_list))
        ws.cell(row=row, column=5, value=", ".join(code_list))

        # sample quotes from related codes
        quotes = _get_quotes_for_theme(theme, codes)
        ws.cell(row=row, column=6, value="; ".join(quotes[:3]))

    # auto-width columns (with max)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 60)

    # wrap text in description and quotes columns
    for row in range(2, len(themes) + 2):
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)

    wb.save(output_path)


def generate_codes_xlsx(
    codes: List[Dict[str, Any]],
    output_path: Path,
) -> None:
    """Export codes to XLSX file.

    Args:
        codes: List of code dicts
        output_path: Path to write XLSX file
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Codes"

    # header
    headers = ["#", "Name", "Description", "Quote Count", "Quotes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # data
    for row, code in enumerate(codes, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=code.get("name", ""))
        ws.cell(row=row, column=3, value=code.get("description", ""))

        quotes = code.get("quotes", [])
        ws.cell(row=row, column=4, value=len(quotes))

        # join quotes
        quote_texts = []
        for quote in quotes:
            text = (
                quote.get("text", str(quote)) if isinstance(quote, dict) else str(quote)
            )
            quote_texts.append(text)
        ws.cell(row=row, column=5, value="\n\n".join(quote_texts))

    # auto-width columns (with max)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    # for multiline, use first line
                    first_line = str(cell.value).split("\n")[0]
                    max_length = max(max_length, len(first_line))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 60)

    # wrap text in description and quotes columns
    for row in range(2, len(codes) + 2):
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)


def _get_quotes_for_theme(theme: Dict, codes: List[Dict]) -> List[str]:
    """Get quotes from codes belonging to a theme."""
    from soak.models.base import compute_code_hash

    code_refs = theme.get(
        "code_hashes", theme.get("code_slugs", theme.get("codes", []))
    )
    resolved_refs = theme.get("resolved_code_refs", [])

    # build set of code identifiers (hashes, slugs, name-slugs)
    code_ids = set(code_refs)
    for ref in resolved_refs:
        if isinstance(ref, dict):
            if "slug" in ref:
                code_ids.add(ref["slug"])
            if "name" in ref:
                code_ids.add(ref["name"].lower().replace(" ", "-"))
            # add hash from resolved ref
            name = ref.get("name", "")
            desc = ref.get("description", "")
            if name and desc:
                code_ids.add(compute_code_hash(name, desc))

    quotes = []
    for code in codes:
        code_slug = code.get("slug", "")
        code_name = code.get("name", "").lower().replace(" ", "-")
        code_hash = compute_code_hash(code.get("name", ""), code.get("description", ""))
        if code_slug in code_ids or code_name in code_ids or code_hash in code_ids:
            for quote in code.get("quotes", [])[:2]:
                text = (
                    quote.get("text", str(quote))
                    if isinstance(quote, dict)
                    else str(quote)
                )
                quotes.append(text)
                if len(quotes) >= 3:
                    return quotes
    return quotes
