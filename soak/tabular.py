"""Tabular mail merge - generate documents from spreadsheet rows using templates.

This module provides functionality to:
1. Parse CSV/XLSX files into rows
2. Validate Jinja2 templates against available columns
3. Generate document content from templates and row data
"""

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from jinja2 import Environment, StrictUndefined, TemplateSyntaxError, UndefinedError

logger = logging.getLogger(__name__)


@dataclass
class TabularParseResult:
    """Result of parsing a tabular file."""

    columns: list[str]
    rows: list[dict[str, str]]
    total_rows: int
    sheet_names: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class GeneratedDocument:
    """A document generated from a template and row data."""

    name: str
    content: str
    row_index: int
    source_file: str
    warnings: list[str] = field(default_factory=list)


def _normalise_column_name(name: str, seen: dict[str, int]) -> str:
    """Normalise column name and handle duplicates."""
    name = str(name).strip()
    if not name:
        name = "Column"

    base_name = name
    if name in seen:
        seen[name] += 1
        name = f"{base_name}_{seen[name]}"
    else:
        seen[name] = 1

    return name


def _try_decode_csv(file_bytes: bytes) -> tuple[str, str]:
    """Try to decode CSV bytes with multiple encodings."""
    encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]

    for encoding in encodings:
        try:
            return file_bytes.decode(encoding), encoding
        except UnicodeDecodeError:
            continue

    raise ValueError(
        "Could not decode file. Tried: UTF-8, Latin-1, Windows-1252. "
        "Please save the file as UTF-8."
    )


def parse_csv(file_path: Path, header_row: int = 1) -> TabularParseResult:
    """Parse a CSV file.

    Args:
        file_path: Path to CSV file
        header_row: Row number containing headers (1-indexed), or 0 for no headers

    Returns:
        TabularParseResult with columns and rows
    """
    warnings = []

    with open(file_path, "rb") as f:
        file_bytes = f.read()

    content, encoding = _try_decode_csv(file_bytes)
    if encoding != "utf-8":
        warnings.append(f"File decoded using {encoding} encoding")

    reader = csv.reader(io.StringIO(content))
    all_rows = list(reader)

    if not all_rows:
        return TabularParseResult(
            columns=[], rows=[], total_rows=0, warnings=["File is empty"]
        )

    if header_row > 0:
        if header_row > len(all_rows):
            raise ValueError(
                f"Header row {header_row} exceeds file length ({len(all_rows)} rows)"
            )
        header_idx = header_row - 1
        raw_headers = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1 :]
    else:
        max_cols = max(len(row) for row in all_rows) if all_rows else 0
        raw_headers = [f"Column {chr(65 + i)}" for i in range(max_cols)]
        data_rows = all_rows

    seen = {}
    columns = [_normalise_column_name(h, seen) for h in raw_headers]

    rows = []
    for row in data_rows:
        row_dict = {}
        for i, col in enumerate(columns):
            value = row[i] if i < len(row) else ""
            row_dict[col] = str(value).strip()
        rows.append(row_dict)

    return TabularParseResult(
        columns=columns, rows=rows, total_rows=len(rows), warnings=warnings
    )


def parse_xlsx(
    file_path: Path,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
) -> TabularParseResult:
    """Parse an XLSX file.

    Args:
        file_path: Path to XLSX file
        sheet_name: Sheet to parse (default: first sheet)
        header_row: Row number containing headers (1-indexed), or 0 for no headers

    Returns:
        TabularParseResult with columns and rows
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX support. Install with: pip install openpyxl"
        )

    warnings = []

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        if "password" in str(e).lower() or "encrypted" in str(e).lower():
            raise ValueError(
                "File is password-protected. Please remove the password and try again."
            )
        raise ValueError(f"Could not open Excel file: {e}")

    sheet_names = workbook.sheetnames

    if sheet_name:
        if sheet_name not in sheet_names:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {', '.join(sheet_names)}"
            )
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    all_rows = []
    for row in sheet.iter_rows(values_only=True):
        all_rows.append([cell if cell is not None else "" for cell in row])

    workbook.close()

    if not all_rows:
        return TabularParseResult(
            columns=[],
            rows=[],
            total_rows=0,
            sheet_names=sheet_names,
            warnings=["Sheet is empty"],
        )

    if header_row > 0:
        if header_row > len(all_rows):
            raise ValueError(
                f"Header row {header_row} exceeds sheet length ({len(all_rows)} rows)"
            )
        header_idx = header_row - 1
        raw_headers = all_rows[header_idx]
        data_rows = all_rows[header_idx + 1 :]
    else:
        max_cols = max(len(row) for row in all_rows) if all_rows else 0
        raw_headers = [f"Column {chr(65 + i)}" for i in range(max_cols)]
        data_rows = all_rows

    seen = {}
    columns = [_normalise_column_name(h, seen) for h in raw_headers]

    rows = []
    for row in data_rows:
        row_dict = {}
        for i, col in enumerate(columns):
            value = row[i] if i < len(row) else ""
            row_dict[col] = str(value).strip()
        rows.append(row_dict)

    return TabularParseResult(
        columns=columns,
        rows=rows,
        total_rows=len(rows),
        sheet_names=sheet_names,
        warnings=warnings,
    )


def parse_tabular_file(
    file_path: Path,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
) -> TabularParseResult:
    """Parse a CSV or XLSX file.

    Args:
        file_path: Path to file
        sheet_name: Sheet to parse (XLSX only)
        header_row: Row number containing headers (1-indexed), or 0 for no headers

    Returns:
        TabularParseResult with columns and rows
    """
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        return parse_csv(file_path, header_row)
    elif suffix in (".xlsx", ".xls"):
        if suffix == ".xls":
            raise ValueError(
                "Legacy .xls format is not supported. Please save as .xlsx."
            )
        return parse_xlsx(file_path, sheet_name, header_row)
    else:
        raise ValueError(f"Unsupported format: {suffix}. Use .csv or .xlsx")


def _get_template_env() -> Environment:
    """Create a Jinja2 environment for template rendering."""
    return Environment(
        undefined=StrictUndefined,
        autoescape=False,
        keep_trailing_newline=True,
    )


def _find_template_variables(template_str: str) -> set[str]:
    """Extract variable names used in a Jinja2 template."""
    env = _get_template_env()

    try:
        parsed = env.parse(template_str)
    except TemplateSyntaxError:
        return set()

    from jinja2 import nodes

    variables = set()

    def visit(node):
        if isinstance(node, nodes.Name):
            variables.add(node.name)
        for child in node.iter_child_nodes():
            visit(child)

    visit(parsed)

    builtins = {"range", "dict", "lipsum", "cycler", "joiner", "namespace", "_columns"}
    return variables - builtins


def validate_template(template_str: str, columns: list[str]) -> list[str]:
    """Validate a Jinja2 template against available columns.

    Returns:
        List of error/warning messages (empty if valid)
    """
    messages = []
    env = _get_template_env()

    try:
        env.parse(template_str)
    except TemplateSyntaxError as e:
        messages.append(f"Template syntax error: {e.message}")
        return messages

    used_vars = _find_template_variables(template_str)
    column_set = set(columns)

    for var in used_vars:
        if var not in column_set:
            similar = [
                c for c in columns if c.lower() == var.lower() or var.lower() in c.lower()
            ]
            if similar:
                messages.append(f"Unknown column '{var}'. Did you mean: {', '.join(similar)}?")
            else:
                messages.append(f"Unknown column '{var}'")

    if not used_vars:
        messages.append("Warning: Template doesn't use any columns. All documents will be identical.")

    return messages


def render_template(template_str: str, row: dict[str, str], columns: list[str]) -> str:
    """Render a template with row data."""
    env = _get_template_env()
    template = env.from_string(template_str)

    context = dict(row)
    context["_columns"] = [row.get(col, "") for col in columns]

    return template.render(**context)


def generate_documents(
    template_str: str,
    parse_result: TabularParseResult,
    source_file: str,
    name_column: Optional[str] = None,
    skip_empty: bool = False,
) -> list[GeneratedDocument]:
    """Generate documents from a template and parsed tabular data.

    Args:
        template_str: The Jinja2 template
        parse_result: Parsed tabular data
        source_file: Original filename for metadata
        name_column: Column to use for document names (default: Row 1, Row 2, ...)
        skip_empty: Skip rows where used columns are all empty

    Returns:
        List of generated documents
    """
    documents = []
    used_columns = _find_template_variables(template_str)

    for i, row in enumerate(parse_result.rows, start=1):
        warnings = []

        empty_used = [col for col in used_columns if col in row and not row[col].strip()]
        if empty_used:
            if skip_empty and all(
                not row.get(col, "").strip() for col in used_columns if col in row
            ):
                continue
            warnings.append(f"Empty values in: {', '.join(empty_used)}")

        if name_column and name_column in row:
            name = row[name_column].strip()
            if not name:
                name = f"Row {i}"
        else:
            name = f"Row {i}"

        try:
            content = render_template(template_str, row, parse_result.columns)
        except UndefinedError as e:
            match = re.search(r"'(\w+)'", str(e))
            var_name = match.group(1) if match else "unknown"
            warnings.append(f"Missing column: {var_name}")
            continue
        except Exception as e:
            warnings.append(f"Template error: {e}")
            continue

        documents.append(
            GeneratedDocument(
                name=name,
                content=content,
                row_index=i,
                source_file=source_file,
                warnings=warnings,
            )
        )

    return documents


def generate_auto_template(columns: list[str]) -> str:
    """Generate a default template with all columns as markdown blockquotes.

    Args:
        columns: List of column names

    Returns:
        Template string with each column as a blockquote section
    """
    lines = []
    for col in columns:
        lines.append(f"{col}:")
        lines.append(f"> {{{{ {col} }}}}")
        lines.append("")
    return "\n".join(lines).strip()


def estimate_total_size(documents: list[GeneratedDocument]) -> int:
    """Estimate total size of generated documents in bytes.

    Args:
        documents: List of generated documents

    Returns:
        Total size in bytes
    """
    return sum(len(doc.content.encode("utf-8")) for doc in documents)
